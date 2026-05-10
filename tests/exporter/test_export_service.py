from io import BytesIO

import pytest
from django.test import override_settings
from openpyxl import load_workbook

from customer.models import Customer, Tag
from exporter.services.export_service import ExportService

pytestmark = pytest.mark.django_db


def _make_customer(*, name, phone, created_by, assigned_to=None, tag=None):
    return Customer.objects.create(
        customer_name=name,
        customer_surname="User",
        customer_phone=phone,
        created_by=created_by,
        assigned_to=assigned_to,
        tag=tag,
    )


def test_export_service_uses_default_fields_and_saves_csv(
    tmp_path,
    admin_user,
    regular_user,
):
    tag = Tag.objects.create(tag_name="VIP", color="#FF0000", description="desc")
    _make_customer(
        name="Alpha",
        phone="1000000100",
        created_by=admin_user,
        assigned_to=regular_user,
        tag=tag,
    )

    with override_settings(MEDIA_ROOT=tmp_path, EXPORT_FILES_ROOT=tmp_path / "exports"):
        result = ExportService().create_export(
            user=admin_user,
            model_name="customer",
            file_type="csv",
        )

    content = result.absolute_path.read_text()
    header = content.splitlines()[0]

    assert result.file_name.endswith(".csv")
    assert result.row_count == 1
    assert result.fields == [
        "id",
        "customer_name",
        "customer_surname",
        "customer_email",
        "customer_phone",
        "status",
        "assigned_to",
        "tag",
        "created_at",
    ]
    assert "customer_name" in header
    assert "assigned_to" in header
    assert result.absolute_path.exists()


def test_export_service_limits_regular_user_rows(tmp_path, admin_user, regular_user):
    other_user = type(regular_user).objects.create_user(
        username="other-exporter",
        email="other-exporter@example.com",
        password="OtherPass123!",
    )
    _make_customer(
        name="Mine",
        phone="1000000101",
        created_by=admin_user,
        assigned_to=regular_user,
    )
    _make_customer(
        name="NotMine",
        phone="1000000102",
        created_by=admin_user,
        assigned_to=other_user,
    )

    with override_settings(MEDIA_ROOT=tmp_path, EXPORT_FILES_ROOT=tmp_path / "exports"):
        result = ExportService().create_export(
            user=regular_user,
            model_name="customer",
            file_type="excel",
            fields=["customer_name", "assigned_to"],
        )

    workbook = load_workbook(filename=BytesIO(result.absolute_path.read_bytes()))
    rows = list(workbook.active.iter_rows(values_only=True))

    assert result.file_name.endswith(".xlsx")
    assert result.row_count == 1
    assert rows[0] == ("customer_name", "assigned_to")
    assert rows[1] == ("Mine", regular_user.username)


def test_export_service_deletes_created_file(tmp_path, admin_user):
    _make_customer(
        name="DeleteMe",
        phone="1000000104",
        created_by=admin_user,
    )

    with override_settings(MEDIA_ROOT=tmp_path, EXPORT_FILES_ROOT=tmp_path / "exports"):
        service = ExportService()
        result = service.create_export(
            user=admin_user,
            model_name="customer",
            file_type="csv",
        )
        deleted = service.delete_export(relative_path=result.relative_path)

    assert deleted is True
    assert result.absolute_path.exists() is False


def test_export_service_delete_returns_false_for_missing_file(tmp_path):
    with override_settings(MEDIA_ROOT=tmp_path, EXPORT_FILES_ROOT=tmp_path / "exports"):
        deleted = ExportService().delete_export(
            relative_path="exports/customer/missing.csv",
        )

    assert deleted is False
