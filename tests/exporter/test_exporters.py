import csv
from datetime import datetime, time, timezone
from io import BytesIO, StringIO

import pytest
from openpyxl import load_workbook

from exporter.exportFactory.csv_exporter import CSVExporter
from exporter.exportFactory.excel_exporter import ExcelExporter
from exporter.exportFactory.factory import ExporterFactory


class TestCSVExporter:
    def test_export_returns_utf8_bytes_with_header_and_rows(self):
        exporter = CSVExporter()
        result = exporter.export(
            headers=["isim", "soyisim"],
            rows=[["Ali", "Veli"], ["Ayşe", "Fatma"]],
        )

        assert isinstance(result, bytes)
        reader = csv.reader(StringIO(result.decode("utf-8")))
        rows = list(reader)
        assert rows[0] == ["isim", "soyisim"]
        assert rows[1] == ["Ali", "Veli"]
        assert rows[2] == ["Ayşe", "Fatma"]

    def test_export_empty_rows_produces_header_only(self):
        exporter = CSVExporter()
        result = exporter.export(headers=["col"], rows=[])

        reader = csv.reader(StringIO(result.decode("utf-8")))
        rows = list(reader)
        assert rows == [["col"]]

    def test_file_type_and_extension(self):
        exporter = CSVExporter()
        assert exporter.file_type == "csv"
        assert exporter.file_extension == "csv"


class TestExcelExporter:
    def _load_rows(self, result: bytes):
        wb = load_workbook(filename=BytesIO(result))
        return list(wb.active.iter_rows(values_only=True))

    def test_export_returns_valid_xlsx_with_header_and_rows(self):
        exporter = ExcelExporter()
        result = exporter.export(
            headers=["isim", "soyisim"],
            rows=[["Ali", "Veli"], ["Ayşe", "Fatma"]],
        )

        rows = self._load_rows(result)
        assert rows[0] == ("isim", "soyisim")
        assert rows[1] == ("Ali", "Veli")
        assert rows[2] == ("Ayşe", "Fatma")

    def test_export_empty_rows_produces_header_only(self):
        exporter = ExcelExporter()
        result = exporter.export(headers=["col"], rows=[])

        rows = self._load_rows(result)
        assert rows == [("col",)]

    def test_worksheet_title_is_export(self):
        exporter = ExcelExporter()
        result = exporter.export(headers=["x"], rows=[])
        wb = load_workbook(filename=BytesIO(result))
        assert wb.active.title == "Export"

    def test_file_type_and_extension(self):
        exporter = ExcelExporter()
        assert exporter.file_type == "excel"
        assert exporter.file_extension == "xlsx"

    def test_normalize_cell_strips_timezone_from_datetime(self):
        dt_aware = datetime(2024, 1, 15, 10, 30, tzinfo=timezone.utc)
        result = ExcelExporter._normalize_cell(dt_aware)
        assert result.tzinfo is None
        assert result == datetime(2024, 1, 15, 10, 30)

    def test_normalize_cell_passes_naive_datetime_unchanged(self):
        dt_naive = datetime(2024, 1, 15, 10, 30)
        result = ExcelExporter._normalize_cell(dt_naive)
        assert result == dt_naive

    def test_normalize_cell_strips_timezone_from_time(self):
        t_aware = time(10, 30, tzinfo=timezone.utc)
        result = ExcelExporter._normalize_cell(t_aware)
        assert result.tzinfo is None
        assert result == time(10, 30)

    def test_normalize_cell_passes_naive_time_unchanged(self):
        t_naive = time(10, 30)
        result = ExcelExporter._normalize_cell(t_naive)
        assert result == t_naive

    def test_normalize_cell_passes_other_types_unchanged(self):
        assert ExcelExporter._normalize_cell("text") == "text"
        assert ExcelExporter._normalize_cell(42) == 42
        assert ExcelExporter._normalize_cell(None) is None


class TestExporterFactory:
    def test_create_returns_csv_exporter(self):
        exporter = ExporterFactory.create("csv")
        assert isinstance(exporter, CSVExporter)

    def test_create_returns_excel_exporter_for_excel(self):
        exporter = ExporterFactory.create("excel")
        assert isinstance(exporter, ExcelExporter)

    def test_create_returns_excel_exporter_for_xlsx_alias(self):
        exporter = ExporterFactory.create("xlsx")
        assert isinstance(exporter, ExcelExporter)

    def test_create_normalizes_case_and_whitespace(self):
        exporter = ExporterFactory.create("  CSV  ")
        assert isinstance(exporter, CSVExporter)

    def test_create_raises_for_unknown_type(self):
        with pytest.raises(ValueError, match="Unsupported file type"):
            ExporterFactory.create("pdf")

    def test_supported_types_contains_csv_and_excel(self):
        types = ExporterFactory.supported_types()
        assert "csv" in types
        assert "excel" in types
